import streamlit as st
import pandas as pd
import io
from collections import Counter
from openpyxl.styles import Font

from animation import render_header_animation
from parser_logic import extract_name_and_prediction, parse_score, get_verdict, process_excel_file, pick_winners

st.set_page_config(page_title="Football Prediction Analyzer", page_icon="⚽", layout="wide", initial_sidebar_state="expanded")

# Render modularized animation
render_header_animation()

st.title("Football Prediction Analyzer")
st.markdown("Easily extract, analyze, and export predictions from social media comments for the World Cup! 🏆")
st.divider()

# Global Match Configuration
st.subheader("⚙️ Match Configuration")
team_col1, team_col2 = st.columns(2)
with team_col1:
    team_1 = st.text_input("First Team Name", value="Brazil")
with team_col2:
    team_2 = st.text_input("Second Team Name", value="Morocco")

# Sidebar
with st.sidebar:
    st.header("📋 About")
    st.info("This tool helps you quickly turn messy social media comment blocks into structured Excel/CSV data.")
    
    with st.expander("📝 Formatting Guide", expanded=True):
        st.markdown("""
        **How to copy comments:**
        Simply highlight the comment section on Facebook/WhatsApp and copy.
        
        **Expected Format:**
        ```text
        Name Surname
        Brazil 2-1 Morocco
        1d
        Reply
        ```
        """)
    st.markdown("---")
    st.caption("Built with Streamlit & ❤️")

def generate_winners_excel_buffer(winners_df, name_col):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        export_df = winners_df.copy()
        if 'Profile Link' in export_df.columns:
            export_df = export_df.drop(columns=['Profile Link'])
        export_df.to_excel(writer, index=False, sheet_name='Winners')
        
        ws = writer.sheets['Winners']
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        if 'Profile Link' in winners_df.columns:
            name_idx = list(export_df.columns).index(name_col) + 1
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                link = winners_df.iloc[row_idx-2].get('Profile Link')
                if pd.notna(link) and link:
                    name_cell = ws.cell(row=row_idx, column=name_idx)
                    name_cell.hyperlink = str(link)
                    name_cell.style = "Hyperlink"
                    
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
    return buffer.getvalue()

def render_winner_picker(df, name_col, text_col, key_prefix, team1, team2):
    """Helper to render the winner picker UI"""
    st.divider()
    st.subheader("🏆 Pick Winners")
    st.markdown("Enter the **actual match score** to find and randomly select people who predicted it exactly.")
    
    w_col1, w_col2, w_col3 = st.columns([1, 1, 1])
    with w_col1:
        actual_t1 = st.number_input(f"Actual {team1} Goals", min_value=0, max_value=20, value=0, step=1, key=f"{key_prefix}_b")
    with w_col2:
        actual_t2 = st.number_input(f"Actual {team2} Goals", min_value=0, max_value=20, value=0, step=1, key=f"{key_prefix}_m")
    with w_col3:
        max_w = st.number_input("Number of Winners", min_value=1, max_value=100, value=10, step=1, key=f"{key_prefix}_w")
        
    if st.button(f"🎲 Randomly Pick {max_w} Winners!", type="primary", key=f"{key_prefix}_btn"):
        winners_df = pick_winners(df, actual_t1, actual_t2, team1, team2, max_winners=max_w)
        if winners_df.empty:
            st.error(f"No one predicted exactly {actual_t1}-{actual_t2}!")
        else:
            st.success(f"Found {len(winners_df)} winners!")
            st.balloons()
            st.dataframe(
                winners_df[[name_col, text_col]], 
                use_container_width=True, 
                hide_index=True
            )
            # Store winners in session state so they can be included in Excel export
            st.session_state[f'{key_prefix}_winners'] = winners_df
            st.session_state[f'{key_prefix}_score'] = f"{actual_t1}-{actual_t2}"

    if f'{key_prefix}_winners' in st.session_state:
        winners_df = st.session_state[f'{key_prefix}_winners']
        actual_score = st.session_state.get(f'{key_prefix}_score', '')
        
        buffer_bytes = generate_winners_excel_buffer(winners_df, name_col)
        
        st.download_button(
            label="📥 Download Winners as Excel (with Links)",
            data=buffer_bytes,
            file_name=f"Winners_{actual_score.replace('-', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_dl_winners_btn"
        )

# Main input area
tab2, tab1 = st.tabs(["📁 Excel File Analysis", "📝 Paste Text Analysis"])

with tab1:
    input_text = st.text_area(
        "📝 Paste your comments here:",
        height=400,
        placeholder="""
Example:
Samiul Islam
My prediction:
Brazil🇧🇷 2-1Morocco🇲🇦
1d
Reply
Shadin Mahbub
Brazil 2-1 Morocco
1d
Reply
    """
    )

    # Analyze button
    if st.button("🔍 Analyze Predictions", type="primary", use_container_width=True):
        if not input_text.strip():
            st.warning("Please paste some comments first!")
        else:
            # Extract data
            raw_results = extract_name_and_prediction(input_text, team_1, team_2)
            
            # Parse scores
            parsed_results = []
            for item in raw_results:
                t1_score, t2_score = parse_score(item['raw_prediction'], team_1, team_2)
                parsed_results.append({
                    'Name': item['name'],
                    'Profile Link': None,
                    'Prediction Text': item['raw_prediction'][:100] if item['raw_prediction'] else 'Not found',
                    f'{team_1} Goals': t1_score if t1_score is not None else '-',
                    f'{team_2} Goals': t2_score if t2_score is not None else '-',
                    'Verdict': get_verdict(t1_score, t2_score, team_1, team_2)
                })
            
            # Filter out entries with no name or completely invalid
            valid_results = [r for r in parsed_results if r['Name'] and r[f'{team_1} Goals'] != '-']
            invalid_results = [r for r in parsed_results if r['Name'] and r[f'{team_1} Goals'] == '-']
            
            # Create DataFrame
            if not valid_results:
                st.warning("No valid predictions could be parsed from the text.")
                if 'text_df' in st.session_state: del st.session_state['text_df']
            else:
                st.session_state['text_df'] = pd.DataFrame(valid_results)
                st.session_state['text_invalid'] = invalid_results
                st.session_state['text_valid'] = valid_results
                st.toast(f"Successfully parsed {len(valid_results)} predictions!", icon="🎉")
                
    if 'text_df' in st.session_state:
        df = st.session_state['text_df']
        valid_results = st.session_state['text_valid']
        invalid_results = st.session_state['text_invalid']
        
        # Display statistics
        st.subheader("📊 Statistics")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Valid Predictions", len(valid_results))
        with col2:
            t1_wins = sum(1 for r in valid_results if r['Verdict'] == f"✅ {team_1} Win")
            st.metric(f"{team_1} Win Predictions", t1_wins)
        with col3:
            t2_wins = sum(1 for r in valid_results if r['Verdict'] == f"❌ {team_2} Win (Upset)")
            st.metric(f"{team_2} Win Predictions", t2_wins)
        with col4:
            draws = sum(1 for r in valid_results if r['Verdict'] == "🤝 Draw")
            st.metric("Draw Predictions", draws)
        
        # Goal distribution charts
        st.subheader("📈 Goal Distribution")
        
        col1, col2 = st.columns(2)
        
        with col1:
            t1_goals_list = [r[f'{team_1} Goals'] for r in valid_results if isinstance(r[f'{team_1} Goals'], int)]
            if t1_goals_list:
                goal_counts = Counter(t1_goals_list)
                st.bar_chart(pd.DataFrame.from_dict(goal_counts, orient='index', columns=['Count']))
        
        with col2:
            total_goals = [r[f'{team_1} Goals'] + r[f'{team_2} Goals'] for r in valid_results if isinstance(r[f'{team_1} Goals'], int) and isinstance(r[f'{team_2} Goals'], int)]
            if total_goals:
                goal_counts = Counter(total_goals)
                st.bar_chart(pd.DataFrame.from_dict(goal_counts, orient='index', columns=['Count']))
        
        # Most common scorelines
        st.subheader("🎯 Most Common Scoreline Predictions")
        scorelines = [f"{r[f'{team_1} Goals']}-{r[f'{team_2} Goals']}" for r in valid_results if isinstance(r[f'{team_1} Goals'], int)]
        if scorelines:
            common = Counter(scorelines).most_common(5)
            for score, count in common:
                st.write(f"**{score}** → {count} predictions")
        
        # Display the main table
        st.subheader("📋 All Predictions")
        display_df = df.drop(columns=['Profile Link']) if 'Profile Link' in df.columns else df
        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Name": st.column_config.TextColumn("Name", width="medium"),
                "Prediction Text": st.column_config.TextColumn("Raw Prediction", width="large"),
                f"{team_1} Goals": st.column_config.NumberColumn(team_1, width="small"),
                f"{team_2} Goals": st.column_config.NumberColumn(team_2, width="small"),
                "Verdict": st.column_config.TextColumn("Verdict", width="medium"),
            }
        )
        
        # Show invalid predictions
        if invalid_results:
            with st.expander("⚠️ Unparsed Comments (needs manual review)"):
                for inv in invalid_results:
                    st.write(f"**{inv['Name']}**: {inv['Prediction Text']}")
        
        # Download buttons
        col1, col2 = st.columns(2)
        
        with col1:
            export_df = df.copy()
            if 'Profile Link' in export_df.columns:
                export_df = export_df.drop(columns=['Profile Link'])
            csv = export_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Download as CSV",
                data=csv,
                file_name="predictions_analysis.csv",
                mime="text/csv",
            )
            
        with col2:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                export_df.to_excel(writer, index=False, sheet_name='Predictions')
                
                worksheet = writer.sheets['Predictions']
                
                # Make headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    
                # Adjust column widths dynamically
                for col in worksheet.columns:
                    max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50) # Cap width at 50

                # Add Winners sheet if winners have been picked
                if 'tab1_winners' in st.session_state:
                    winners_df = st.session_state['tab1_winners']
                    export_winners_df = winners_df.copy()
                    if 'Profile Link' in export_winners_df.columns:
                        export_winners_df = export_winners_df.drop(columns=['Profile Link'])
                    export_winners_df.to_excel(writer, index=False, sheet_name='Winners')
                    
                    ws_winners = writer.sheets['Winners']
                    for cell in ws_winners[1]:
                        cell.font = Font(bold=True)
                    for col in ws_winners.columns:
                        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
                        col_letter = col[0].column_letter
                        ws_winners.column_dimensions[col_letter].width = min(max_length + 2, 50)

            st.download_button(
                label="📊 Download as Excel",
                data=buffer.getvalue(),
                file_name="predictions_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            
        render_winner_picker(df, 'Name', 'Prediction Text', 'tab1', team_1, team_2)

with tab2:
    st.markdown("### 📁 Upload Raw Comments Excel File")
    st.info("Upload a raw single-column Excel file containing comments (like `Sample.xlsx`) to parse it into structured columns with working hyperlinks.")
    uploaded_file = st.file_uploader("Choose an Excel file (.xlsx)", type=['xlsx'])
    
    if uploaded_file is not None:
        if st.button("⚙️ Process Excel File", type="primary", use_container_width=True):
            with st.spinner("Processing your file..."):
                try:
                    output_buffer, preview_df = process_excel_file(uploaded_file, team_1, team_2)
                    st.session_state['excel_output'] = output_buffer
                    st.session_state['excel_df'] = preview_df
                    st.toast("File processed successfully!", icon="✅")
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")
                    
        if 'excel_df' in st.session_state:
            preview_df = st.session_state['excel_df']
            output_buffer = st.session_state['excel_output']
            
            st.subheader("📊 Extraction Summary")
            st.metric("Total Comments Extracted", len(preview_df))
            
            st.subheader("Preview (first 10 rows)")
            display_preview = preview_df.drop(columns=['Profile Link']) if 'Profile Link' in preview_df.columns else preview_df
            st.dataframe(display_preview.head(10), use_container_width=True, hide_index=True)
            
            # Build Excel with optional Winners tab preserving original output
            import openpyxl
            output_buffer.seek(0)
            wb = openpyxl.load_workbook(output_buffer)
            
            # Add Winners sheet if winners have been picked
            if 'tab2_winners' in st.session_state:
                if 'Winners' in wb.sheetnames:
                    del wb['Winners']
                    
                ws_winners = wb.create_sheet('Winners')
                winners_df = st.session_state['tab2_winners']
                
                export_cols = [c for c in winners_df.columns if c != 'Profile Link']
                ws_winners.append(export_cols)
                for cell in ws_winners[1]:
                    cell.font = Font(bold=True)
                
                if 'Commenter Name' in export_cols:
                    name_idx = export_cols.index('Commenter Name') + 1
                else:
                    name_idx = 1
                    
                for r_idx, row in enumerate(winners_df.to_dict('records'), start=2):
                    row_values = [row[c] for c in export_cols]
                    ws_winners.append(row_values)
                    
                    link = row.get('Profile Link')
                    if pd.notna(link) and link:
                        name_cell = ws_winners.cell(row=r_idx, column=name_idx)
                        name_cell.hyperlink = str(link)
                        name_cell.style = "Hyperlink"
                        
                for col in ws_winners.columns:
                    max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
                    col_letter = col[0].column_letter
                    ws_winners.column_dimensions[col_letter].width = min(max_length + 2, 50)

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="📥 Download Analyzed Excel File",
                data=final_buffer.getvalue(),
                file_name="Analyzed_" + uploaded_file.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            render_winner_picker(preview_df, 'Commenter Name', 'Comment', 'tab2', team_1, team_2)