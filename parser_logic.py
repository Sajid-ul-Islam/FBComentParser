import re
import io
import openpyxl
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font

def extract_name_and_prediction(text, team1="Brazil", team2="Morocco"):
    """Extract name and prediction from comment blocks"""
    lines = text.strip().split('\n')
    results = []
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Skip empty lines
        if not line:
            i += 1
            continue
        
        # Check if line looks like a name (not containing numbers or common prediction patterns)
        is_name = (
            len(line) > 2 and 
            not re.search(r'\d+[dmhyw]$', line) and # Exclude timestamps like 1d, 2h, 5m
            not re.search(rf'{team1}|{team2}|ব্রাজিল|মরক্কো', line, re.IGNORECASE) and
            not re.search(r'goal|গোল|prediction', line, re.IGNORECASE) and
            line not in ['Reply', 'Edited', 'See Original', 'Top fan', 'Follow', 'Like'] and
            not re.search(r'^[\W_]+$', line) # Exclude lines with only emojis/punctuation
        )
        
        if is_name:
            name = line
            # Look ahead for prediction
            prediction = ""
            lines_to_skip = 0
            for j in range(i+1, min(i+5, len(lines))):
                next_line = lines[j].strip()
                # Check if this line contains a score prediction
                if re.search(r'\d+\s*[-–:]\s*\d+', next_line):
                    prediction = next_line
                    lines_to_skip = j - i
                    break
                elif re.search(rf'{team1}.*\d.*{team2}|ব্রাজিল.*\d.*মরক্কো', next_line, re.IGNORECASE):
                    prediction = next_line
                    lines_to_skip = j - i
                    break
            
            # If no prediction found in next lines, check current line
            if not prediction and re.search(r'\d+\s*[-–:]\s*\d+', line):
                prediction = line
            
            results.append({
                'name': name,
                'raw_prediction': prediction
            })
            if lines_to_skip > 0:
                i += lines_to_skip + 1
            else:
                i += 1
        else:
            i += 1
    
    return results

def parse_score(prediction, team1="Brazil", team2="Morocco"):
    """Extract Brazil and Morocco scores from prediction text"""
    if not prediction:
        return None, None
    
    # Pattern for "Brazil 2-1 Morocco" or "ব্রাজিল ২-১ মরক্কো"
    patterns = [
        rf'{team1}[^\d]*(\d+)\s*[-–:]\s*(\d+)[^\d]*{team2}',
        r'ব্রাজিল[^\d]*(\d+)\s*[-–:]\s*(\d+)[^\d]*মরক্কো',
        r'(\d+)\s*[-–:]\s*(\d+)',  # Just numbers like "2-1"
    ]
    
    for pattern in patterns:
        match = re.search(pattern, prediction, re.IGNORECASE)
        if match:
            return int(match.group(1)), int(match.group(2))
    
    return None, None

def get_verdict(brazil_score, morocco_score, team1="Brazil", team2="Morocco"):
    """Generate verdict based on prediction"""
    if brazil_score is None or morocco_score is None:
        return "❓ Invalid Prediction"
    
    if brazil_score > morocco_score:
        return f"✅ {team1} Win"
    elif morocco_score > brazil_score:
        return f"❌ {team2} Win (Upset)"
    else:
        return "🤝 Draw"

def process_excel_file(uploaded_file, team1="Brazil", team2="Morocco"):
    """Parse raw excel file and extract formatted commenter info"""
    wb = openpyxl.load_workbook(uploaded_file)
    sheet = wb.active
    
    out_wb = openpyxl.Workbook()
    out_sheet = out_wb.active
    out_sheet.append(["Commenter Name", "Comment", "Time Ago", f"{team1} Goals", f"{team2} Goals", "Verdict"])
    
    # Make header bold
    for cell in out_sheet[1]:
        cell.font = Font(bold=True)
        
    # Adjust column widths
    out_sheet.column_dimensions['A'].width = 25  # Commenter Name
    out_sheet.column_dimensions['B'].width = 60  # Comment
    out_sheet.column_dimensions['C'].width = 12  # Time Ago
    out_sheet.column_dimensions['D'].width = 15  # Team 1 Goals
    out_sheet.column_dimensions['E'].width = 15  # Team 2 Goals
    out_sheet.column_dimensions['F'].width = 25  # Verdict

    current_name = None
    current_link = None
    current_comment_lines = []
    preview_data = []
    
    for row in sheet.iter_rows():
        if not row: continue
        cell = row[0]
        val = cell.value
        if val is None: continue
        
        val = str(val).strip()
        if not val: continue
        
        link = cell.hyperlink.target if cell.hyperlink else None
        
        if re.match(r'^\d+\s*[dmhyws]$', val, re.IGNORECASE) or val.lower() == 'just now':
            time_ago = val
            if current_name:
                comment_text = "\\n".join(current_comment_lines)
                brazil, morocco = parse_score(comment_text, team1, team2)
                verdict = get_verdict(brazil, morocco, team1, team2)
                
                row_idx = out_sheet.max_row + 1
                name_cell = out_sheet.cell(row=row_idx, column=1, value=current_name)
                if current_link:
                    name_cell.hyperlink = current_link
                    name_cell.style = "Hyperlink"
                out_sheet.cell(row=row_idx, column=2, value=ILLEGAL_CHARACTERS_RE.sub(r'', comment_text))
                out_sheet.cell(row=row_idx, column=3, value=time_ago)
                out_sheet.cell(row=row_idx, column=4, value=brazil if brazil is not None else '-')
                out_sheet.cell(row=row_idx, column=5, value=morocco if morocco is not None else '-')
                out_sheet.cell(row=row_idx, column=6, value=verdict)
                
                # Append to preview data
                preview_data.append({
                    "Commenter Name": current_name,
                    "Profile Link": current_link,
                    "Comment": ILLEGAL_CHARACTERS_RE.sub(r'', comment_text),
                    "Time Ago": time_ago,
                    f"{team1} Goals": brazil if brazil is not None else '-',
                    f"{team2} Goals": morocco if morocco is not None else '-',
                    "Verdict": verdict
                })
                
            current_name = None
            current_link = None
            current_comment_lines = []
        elif val in ['Reply', 'Hide', 'Send message', 'See Translation', 'Edited', 'Top fan', 'Follow', 'Like']:
            continue
        else:
            if not current_name:
                current_name = val
                current_link = link
            else:
                current_comment_lines.append(val)
                
    output = io.BytesIO()
    out_wb.save(output)
    output.seek(0)
    
    preview_df = pd.DataFrame(preview_data)
    
    return output, preview_df

def pick_winners(df, actual_brazil, actual_morocco, team1="Brazil", team2="Morocco", max_winners=10):
    """Filter for exact score predictions and randomly select up to max_winners"""
    if df.empty:
        return pd.DataFrame()
        
    # Clean df to only include rows where goals are numeric (handles the '-' placeholder)
    df_clean = df[pd.to_numeric(df[f'{team1} Goals'], errors='coerce').notnull()]
    
    if df_clean.empty:
        return pd.DataFrame()
        
    # Filter for exact score match
    correct_df = df_clean[
        (df_clean[f'{team1} Goals'] == actual_brazil) & 
        (df_clean[f'{team2} Goals'] == actual_morocco)
    ]
    
    if correct_df.empty:
        return correct_df
        
    # Pick random winners
    winners = correct_df.sample(n=min(len(correct_df), max_winners))
    return winners
